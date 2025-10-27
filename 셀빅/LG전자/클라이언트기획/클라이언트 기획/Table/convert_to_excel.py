import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# CSV 파일 읽기
csv_path = r'c:\Users\user\Desktop\송형동_아카이브\셀빅\LG전자\클라이언트기획\클라이언트 기획\Table\LG_클라이언트_테이블_통합.csv'
xlsx_path = r'c:\Users\user\Desktop\송형동_아카이브\셀빅\LG전자\클라이언트기획\클라이언트 기획\Table\LG_클라이언트_테이블_통합.xlsx'

# 워크북 생성
wb = Workbook()
wb.remove(wb.active)  # 기본 시트 제거

# 스타일 정의
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
section_font = Font(bold=True, size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# CSV 파일 읽기
with open(csv_path, 'r', encoding='utf-8') as f:
    lines = f.readlines()

# 시트별로 데이터 분리
current_sheet = None
current_data = []
sheet_data = {}

for line in lines:
    line = line.strip()
    if not line:
        continue
    
    # 섹션 구분
    if line.startswith('==================== 1. CONTENTS'):
        if current_sheet and current_data:
            sheet_data[current_sheet] = current_data
        current_sheet = 'CONTENTS'
        current_data = []
    elif line.startswith('==================== 2. THEMES'):
        if current_sheet and current_data:
            sheet_data[current_sheet] = current_data
        current_sheet = 'THEMES'
        current_data = []
    elif line.startswith('==================== 3. EXTERNAL_PARAMS'):
        if current_sheet and current_data:
            sheet_data[current_sheet] = current_data
        current_sheet = 'EXTERNAL_PARAMS'
        current_data = []
    elif line.startswith('==================== 콘텐츠별'):
        if current_sheet and current_data:
            sheet_data[current_sheet] = current_data
        current_sheet = 'External_Param_조합'
        current_data = []
    elif line.startswith('==================== 클라이언트 API'):
        if current_sheet and current_data:
            sheet_data[current_sheet] = current_data
        current_sheet = 'API_연동'
        current_data = []
    elif line.startswith('==================== 기본값 사용'):
        if current_sheet and current_data:
            sheet_data[current_sheet] = current_data
        current_sheet = '기본값_시나리오'
        current_data = []
    elif not line.startswith('===================='):
        current_data.append(line)

# 마지막 시트 저장
if current_sheet and current_data:
    sheet_data[current_sheet] = current_data

# 1. 개요 시트 생성
ws_overview = wb.create_sheet("📋 개요")
overview_data = [
    ["LG AI 영상 생성 서비스 - 클라이언트 테이블 명세"],
    [""],
    ["테이블명", "테이블 설명", "비고"],
    ["contents", "콘텐츠 정보 (10개)", "클라이언트에서 콘텐츠 목록 조회 시 사용"],
    ["themes", "테마 정보 (16개)", "클라이언트에서 테마 목록 조회 시 사용"],
    ["external_params", "External_Param 정의 (4개)", "클라이언트에서 외부변수 설정 시 사용"],
]

for row_idx, row_data in enumerate(overview_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_overview.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14, color="4472C4")
        elif row_idx == 3:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        elif row_idx > 3:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

# 열 너비 조정
ws_overview.column_dimensions['A'].width = 20
ws_overview.column_dimensions['B'].width = 30
ws_overview.column_dimensions['C'].width = 50

# 2. CONTENTS 시트
ws_contents = wb.create_sheet("📦 CONTENTS")
contents_structure = [
    ["순번", "컬럼명", "데이터타입", "제약조건", "기본값", "설명"],
    [1, "content_id", "INT", "PK, AUTO_INCREMENT", "-", "콘텐츠 고유 번호"],
    [2, "content_code", "VARCHAR(50)", "UNIQUE, NOT NULL", "-", "콘텐츠 영문 코드"],
    [3, "content_name", "VARCHAR(100)", "NOT NULL", "-", "콘텐츠 한글 이름"],
    [4, "description", "TEXT", "-", "NULL", "콘텐츠 상세 설명"],
    [5, "icon", "VARCHAR(10)", "-", "NULL", "아이콘"],
    [6, "estimated_time", "VARCHAR(20)", "-", "NULL", "예상 소요 시간"],
    [7, "clip_count", "INT", "-", "1", "기본 클립 개수"],
    [8, "requires_file_upload", "BOOLEAN", "-", "FALSE", "파일 업로드 필요 여부"],
    [9, "file_types", "JSON", "-", "NULL", "허용 파일 타입 배열"],
    [10, "display_order", "INT", "-", "0", "화면 표시 순서"],
    [11, "is_active", "BOOLEAN", "-", "TRUE", "활성화 상태"],
    [12, "created_at", "TIMESTAMP", "-", "NOW()", "등록 일시"],
    [13, "updated_at", "TIMESTAMP", "-", "NULL", "수정 일시"],
]

for row_idx, row_data in enumerate(contents_structure, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_contents.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

# 데이터 추가
ws_contents.cell(row=15, column=1, value="[데이터 목록]").font = section_font
contents_data = [
    ["content_id", "content_code", "content_name", "description", "icon", "estimated_time", "clip_count", "requires_file_upload", "file_types", "display_order", "is_active"],
    [1, "meditation", "명상", "휴식을 취하면서 감상할 수 있는 힐링 콘텐츠", "🧘", "5-10분", 1, "FALSE", "NULL", 1, "TRUE"],
    [2, "visual_art", "비주얼아트", "운동/공부 등 특정 테마에 맞춘 음악과 영상의 조화", "🎨", "15-20분", 16, "FALSE", "NULL", 2, "TRUE"],
    [3, "pet", "애완동물", "동물을 위한 콘텐츠", "🐕", "10-15분", 2, "FALSE", "NULL", 3, "TRUE"],
    [4, "clock", "시계", "선택한 숫자 폰트와 테마로 시각화한 시계", "⏰", "5분", 1, "FALSE", "NULL", 4, "TRUE"],
    [5, "avatar", "아바타", "인물 사진 기반 캐릭터 행동 콘텐츠", "👤", "20-30분", 1, "TRUE", '["image/jpeg", "image/png"]', 5, "TRUE"],
    [6, "interior", "인테리어", "실내를 색다르게 변경한 콘텐츠", "🏠", "15-20분", 1, "FALSE", "NULL", 6, "TRUE"],
    [7, "memorial_album", "기념앨범", "단체 사진 기반 기념 콘텐츠", "📸", "20-30분", 1, "TRUE", '["image/jpeg", "image/png"]', 7, "TRUE"],
    [8, "music", "음악감상", "사운드 파일 기반 뮤직 비디오", "🎵", "25-35분", 6, "TRUE", '["audio/mpeg", "audio/mp3"]', 8, "TRUE"],
    [9, "weather", "날씨", "날씨와 계절 정보 시각화 콘텐츠", "☀️", "10-15분", 8, "FALSE", "NULL", 9, "TRUE"],
    [10, "plant", "식물키우기", "습도에 따른 식물 상태 관리", "🌱", "10-15분", 3, "FALSE", "NULL", 10, "TRUE"],
]

for row_idx, row_data in enumerate(contents_data, 16):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_contents.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 16:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

# 열 너비 조정
ws_contents.column_dimensions['A'].width = 12
ws_contents.column_dimensions['B'].width = 20
ws_contents.column_dimensions['C'].width = 18
ws_contents.column_dimensions['D'].width = 40
ws_contents.column_dimensions['E'].width = 10
ws_contents.column_dimensions['F'].width = 15

# 3. THEMES 시트
ws_themes = wb.create_sheet("🎨 THEMES")
themes_structure = [
    ["순번", "컬럼명", "데이터타입", "제약조건", "기본값", "설명"],
    [1, "theme_id", "INT", "PK, AUTO_INCREMENT", "-", "테마 고유 번호"],
    [2, "theme_code", "VARCHAR(50)", "UNIQUE, NOT NULL", "-", "테마 영문 코드"],
    [3, "theme_name", "VARCHAR(100)", "NOT NULL", "-", "테마 한글 이름"],
    [4, "content_code", "VARCHAR(50)", "FK → contents(content_code)", "NULL", "연결된 콘텐츠 코드"],
    [5, "description", "TEXT", "-", "NULL", "테마 상세 설명"],
    [6, "display_order", "INT", "-", "0", "화면 표시 순서"],
    [7, "is_active", "BOOLEAN", "-", "TRUE", "활성화 상태"],
    [8, "created_at", "TIMESTAMP", "-", "NOW()", "등록 일시"],
]

for row_idx, row_data in enumerate(themes_structure, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_themes.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

# 데이터 추가
ws_themes.cell(row=10, column=1, value="[비주얼아트 테마 - 7개]").font = section_font
themes_data_visual = [
    ["theme_id", "theme_code", "theme_name", "content_code", "description", "display_order", "is_active"],
    [1, "exercise", "운동", "visual_art", "에너지 넘치고 활기찬 분위기", 1, "TRUE"],
    [2, "study", "공부", "visual_art", "집중력 향상에 도움되는 차분한 분위기", 2, "TRUE"],
    [3, "yoga", "요가", "visual_art", "마음을 안정시키고 이완시키는 편안한 분위기", 3, "TRUE"],
    [4, "walk", "산책", "visual_art", "자연과 함께하는 상쾌하고 평화로운 분위기", 4, "TRUE"],
    [5, "travel", "여행", "visual_art", "새로운 장소와 문화 탐험의 설레는 분위기", 5, "TRUE"],
    [6, "reading", "독서", "visual_art", "조용하고 아늑한 분위기", 6, "TRUE"],
    [7, "driving", "드라이브", "visual_art", "자유롭고 개방된 느낌", 7, "TRUE"],
]

for row_idx, row_data in enumerate(themes_data_visual, 11):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_themes.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 11:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

ws_themes.cell(row=19, column=1, value="[애완동물 테마 - 4개]").font = section_font
themes_data_pet = [
    ["theme_id", "theme_code", "theme_name", "content_code", "description", "display_order", "is_active"],
    [8, "dog", "강아지", "pet", "활발하고 귀여운 행동", 1, "TRUE"],
    [9, "cat", "고양이", "pet", "우아하고 신비로운 행동", 2, "TRUE"],
    [10, "rabbit", "토끼", "pet", "귀엽고 사랑스러운 행동", 3, "TRUE"],
    [11, "hamster", "햄스터", "pet", "작고 앙증맞은 행동", 4, "TRUE"],
]

for row_idx, row_data in enumerate(themes_data_pet, 20):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_themes.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 20:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

ws_themes.cell(row=25, column=1, value="[시계 테마 - 5개]").font = section_font
themes_data_clock = [
    ["theme_id", "theme_code", "theme_name", "content_code", "description", "display_order", "is_active"],
    [12, "nature_clock", "자연", "clock", "자연 테마 시계", 1, "TRUE"],
    [13, "city_clock", "도시", "clock", "도시 테마 시계", 2, "TRUE"],
    [14, "space_clock", "우주", "clock", "우주 테마 시계", 3, "TRUE"],
    [15, "minimal_clock", "미니멀", "clock", "미니멀 테마 시계", 4, "TRUE"],
    [16, "vintage_clock", "빈티지", "clock", "빈티지 테마 시계", 5, "TRUE"],
]

for row_idx, row_data in enumerate(themes_data_clock, 26):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_themes.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 26:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

ws_themes.column_dimensions['A'].width = 12
ws_themes.column_dimensions['B'].width = 18
ws_themes.column_dimensions['C'].width = 15
ws_themes.column_dimensions['D'].width = 18
ws_themes.column_dimensions['E'].width = 40

# 4. EXTERNAL_PARAMS 시트
ws_params = wb.create_sheet("⚙️ EXTERNAL_PARAMS")
params_structure = [
    ["순번", "컬럼명", "데이터타입", "제약조건", "기본값", "설명"],
    [1, "param_id", "INT", "PK, AUTO_INCREMENT", "-", "External_Param 고유 번호"],
    [2, "param_code", "VARCHAR(50)", "UNIQUE, NOT NULL", "-", "External_Param 영문 코드"],
    [3, "param_name", "VARCHAR(100)", "NOT NULL", "-", "External_Param 한글 이름"],
    [4, "param_type", "ENUM", "NOT NULL", "-", "파라미터 타입"],
    [5, "possible_values", "JSON", "-", "NULL", "가능한 값 배열"],
    [6, "value_labels", "JSON", "-", "NULL", "값별 한글 라벨"],
    [7, "default_value", "VARCHAR(50)", "-", "NULL", "기본값 (센서 오류 시)"],
    [8, "auto_detect", "BOOLEAN", "-", "FALSE", "자동 감지 여부"],
    [9, "data_source", "VARCHAR(100)", "-", "NULL", "데이터 소스"],
    [10, "description", "TEXT", "-", "NULL", "External_Param 상세 설명"],
    [11, "display_order", "INT", "-", "0", "화면 표시 순서"],
    [12, "is_active", "BOOLEAN", "-", "TRUE", "활성화 상태"],
    [13, "created_at", "TIMESTAMP", "-", "NOW()", "등록 일시"],
]

for row_idx, row_data in enumerate(params_structure, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_params.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

# 데이터 추가
ws_params.cell(row=15, column=1, value="[데이터 목록 - 4개]").font = section_font
params_data = [
    ["param_id", "param_code", "param_name", "param_type", "default_value", "auto_detect", "data_source", "description"],
    [1, "time_brightness", "시간대/조도", "enum", "day", "TRUE", "system_time", "시간대별 조도 변화"],
    [2, "weather", "날씨", "enum", "sunny", "TRUE", "weather_api", "현재 날씨 상태"],
    [3, "holiday", "기념일", "enum", "weekday", "TRUE", "calendar_api", "특별한 날짜 및 기념일"],
    [4, "humidity", "습도", "enum", "moderate", "TRUE", "iot_sensor", "실내 습도 수준"],
]

for row_idx, row_data in enumerate(params_data, 16):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_params.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 16:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

# 값 상세 설명
ws_params.cell(row=22, column=1, value="[시간대/조도 값]").font = section_font
time_values = [
    ["값코드", "한글명", "설명", "기본값"],
    ["sunrise", "일출", "해가 뜨는 시간대", ""],
    ["day", "낮", "해가 떠 있는 시간대", "✅"],
    ["sunset", "일몰", "해가 지는 시간대", ""],
    ["night", "밤", "해가 진 후 시간대", ""],
]

for row_idx, row_data in enumerate(time_values, 23):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_params.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 23:
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.font = Font(bold=True)
        cell.border = border

ws_params.cell(row=28, column=1, value="[날씨 값]").font = section_font
weather_values = [
    ["값코드", "한글명", "설명", "기본값"],
    ["sunny", "맑음", "맑은 날씨", "✅"],
    ["cloudy", "흐림", "구름 낀 날씨", ""],
    ["rainy", "비", "비가 오는 날씨", ""],
    ["snowy", "눈", "눈이 오는 날씨", ""],
]

for row_idx, row_data in enumerate(weather_values, 29):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_params.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 29:
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.font = Font(bold=True)
        cell.border = border

ws_params.column_dimensions['A'].width = 18
ws_params.column_dimensions['B'].width = 20
ws_params.column_dimensions['C'].width = 18
ws_params.column_dimensions['D'].width = 40

# 5. 조합 시트
ws_combination = wb.create_sheet("🔗 External_Param_조합")
combination_data = [
    ["콘텐츠", "사용 External_Param", "클립 개수", "설명"],
    ["명상", "time_brightness(4) × humidity(3)", "12개", "시간대와 습도 조합"],
    ["비주얼아트", "time_brightness(4) × weather(4)", "16개", "시간대와 날씨 조합"],
    ["날씨", "weather(4) × time_brightness(4)", "16개", "날씨와 시간대 조합"],
    ["식물키우기", "humidity(3)", "3개", "습도만 사용"],
    ["기념일 콘텐츠", "holiday(6)", "6개", "기념일만 사용"],
    ["기타", "없음", "1개", "External_Param 미사용"],
]

for row_idx, row_data in enumerate(combination_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_combination.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

ws_combination.column_dimensions['A'].width = 18
ws_combination.column_dimensions['B'].width = 35
ws_combination.column_dimensions['C'].width = 12
ws_combination.column_dimensions['D'].width = 30

# 워크북 저장
wb.save(xlsx_path)
print(f"✅ 엑셀 파일 생성 완료: {xlsx_path}")
