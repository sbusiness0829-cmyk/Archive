from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 파일명 구성
TITLE = "M22_펭귄 어드벤처_사운드리스트"
VERSION = "v.1.0.0"
DATE_STR = datetime.now().strftime("%y%m%d")  # 예: 251027
FILENAME = f"{TITLE}_{VERSION}_{DATE_STR}.xlsx"

# 공통 스타일
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 열 너비 헬퍼
COL_WIDTHS = {
    "사운드리스트": [5, 12, 14, 18, 26, 28, 10, 8, 16, 36, 10, 10, 12, 10, 12, 12, 24, 12],
    "음성 스크립트": [5, 12, 14, 50, 12, 24, 10, 28, 24],
    "ref": [14, 60, 30],
}

# 시트 헤더 정의
SHEET_HEADERS = {
    "사운드리스트": [
        "No", "Category", "SubCategory", "CueID", "EventName", "FileName",
        "Length(s)", "Loop", "Trigger", "Description", "Priority", "Volume",
        "Format", "Channels", "SampleRate", "Owner", "Notes", "Status"
    ],
    "음성 스크립트": [
        "No", "Character", "LineID", "Text", "Emotion", "Context",
        "Est.Length(s)", "FileName", "Notes"
    ],
    "ref": ["Category", "Reference", "Notes"],
}

# 샘플 데이터 (기획서 기반 가정값, 초안)
SFX_ROWS = [
    # BGM
    (1, "BGM", "Title", "BGM_M22_TITLE", "Play_M22_BGM_Title", "bgm_title_loop.wav", 120, "Y", "OnEnter:Title", "타이틀 화면 배경음악(밝고 경쾌)", 2, -6, "wav", "stereo", 48000, "Audio", "Loop crossfade 1s", "draft"),
    (2, "BGM", "Stage_Snow", "BGM_M22_SNOW", "Play_M22_BGM_Snow", "bgm_stage_snow_loop.wav", 180, "Y", "OnEnter:Stage_Snow", "설원 스테이지 BGM(차분/설렘)", 2, -8, "wav", "stereo", 48000, "Audio", "Loop seamless", "draft"),
    (3, "BGM", "Cave", "BGM_M22_CAVE", "Play_M22_BGM_Cave", "bgm_cave_loop.wav", 150, "Y", "OnEnter:Cave", "얼음 동굴 스테이지 BGM(잔향감)", 3, -8, "wav", "stereo", 48000, "Audio", "Add subtle pad", "draft"),
    (4, "BGM", "Boss", "BGM_M22_BOSS", "Play_M22_BGM_Boss", "bgm_boss_loop.wav", 140, "Y", "OnEnter:Boss", "보스전 BGM(긴장감)", 1, -5, "wav", "stereo", 48000, "Audio", "Sidechain with SFX", "draft"),
    (5, "BGM", "Result_Win", "BGM_M22_WIN", "Play_M22_BGM_Win", "bgm_result_win.wav", 8, "N", "OnResult:Win", "결과 화면 승리 짧은 팬페어", 3, -3, "wav", "stereo", 48000, "Audio", "One-shot", "draft"),
    (6, "BGM", "Result_Lose", "BGM_M22_LOSE", "Play_M22_BGM_Lose", "bgm_result_lose.wav", 6, "N", "OnResult:Lose", "결과 화면 패배 짧은 브릿지", 3, -6, "wav", "stereo", 48000, "Audio", "One-shot", "draft"),

    # Ambience
    (7, "AMB", "Wind_Blizzard", "AMB_M22_WIND", "Play_M22_Amb_Wind", "amb_wind_blizzard_loop.wav", 60, "Y", "OnEnter:Outdoor_Snow", "설원 바람/눈보라", 4, -12, "wav", "stereo", 48000, "Audio", "Loop; random gust layer", "draft"),
    (8, "AMB", "Ocean", "AMB_M22_OCEAN", "Play_M22_Amb_Ocean", "amb_ocean_loop.wav", 60, "Y", "OnEnter:Coast", "바다 파도 소리", 4, -12, "wav", "stereo", 48000, "Audio", "Loop; gentle", "draft"),
    (9, "AMB", "Cave_Drops", "AMB_M22_CAVE_DRIP", "Play_M22_Amb_Cave", "amb_cave_drops_loop.wav", 60, "Y", "OnEnter:Cave", "동굴 물방울/잔향", 4, -14, "wav", "stereo", 48000, "Audio", "Loop; sparse", "draft"),

    # UI
    (10, "UI", "Button", "UI_M22_BTN", "Play_UI_Button", "ui_button.wav", 0.25, "N", "OnClick", "버튼 클릭", 5, -6, "wav", "stereo", 44100, "UI", "Short", "final"),
    (11, "UI", "Confirm", "UI_M22_CONFIRM", "Play_UI_Confirm", "ui_confirm.wav", 0.5, "N", "OnConfirm", "확인/완료", 5, -6, "wav", "stereo", 44100, "UI", "Short", "final"),
    (12, "UI", "Cancel", "UI_M22_CANCEL", "Play_UI_Cancel", "ui_cancel.wav", 0.4, "N", "OnCancel", "취소/닫기", 5, -8, "wav", "stereo", 44100, "UI", "Short", "final"),
    (13, "UI", "Reward", "UI_M22_REWARD", "Play_UI_Reward", "ui_reward.wav", 1.2, "N", "OnReward", "보상 팝업", 4, -6, "wav", "stereo", 44100, "UI", "Sparkle tail", "draft"),
    (14, "UI", "Error", "UI_M22_ERROR", "Play_UI_Error", "ui_error.wav", 0.5, "N", "OnError", "에러/불가", 4, -6, "wav", "stereo", 44100, "UI", "Short", "final"),

    # Player
    (15, "Player", "Footstep_Snow", "PL_M22_STEP_SNOW", "Play_Step_Snow", "pl_step_snow_var.wav", 0.35, "N", "OnMove", "눈길 발자국(버전 4종)", 4, -8, "wav", "stereo", 44100, "SFX", "Randomize pitch 4%", "draft"),
    (16, "Player", "Slide_Ice", "PL_M22_SLIDE", "Play_Slide_Ice", "pl_slide_ice_loop.wav", 2.0, "Y", "OnSlide", "얼음 미끄러짐 루프", 3, -10, "wav", "stereo", 44100, "SFX", "Fade in/out 100ms", "draft"),
    (17, "Player", "Jump", "PL_M22_JUMP", "Play_Jump", "pl_jump.wav", 0.35, "N", "OnJump", "점프", 4, -6, "wav", "stereo", 44100, "SFX", "", "final"),
    (18, "Player", "Land", "PL_M22_LAND", "Play_Land_Snow", "pl_land_snow.wav", 0.45, "N", "OnLand", "착지(눈)", 4, -6, "wav", "stereo", 44100, "SFX", "", "final"),
    (19, "Player", "Collect_Fish", "PL_M22_GET_FISH", "Play_Get_Fish", "pl_collect_fish.wav", 0.6, "N", "OnCollect:Fish", "먹이(물고기) 획득", 4, -6, "wav", "stereo", 44100, "SFX", "Shimmer", "draft"),
    (20, "Player", "Damage", "PL_M22_HIT", "Play_Player_Hit", "pl_hit.wav", 0.4, "N", "OnHit", "피격", 3, -6, "wav", "stereo", 44100, "SFX", "", "final"),
    (21, "Player", "Death", "PL_M22_DEATH", "Play_Player_Death", "pl_death.wav", 1.0, "N", "OnDeath", "사망", 2, -6, "wav", "stereo", 44100, "SFX", "", "draft"),

    # Enemy / Object
    (22, "Enemy", "Seal_Roar", "EN_M22_SEAL", "Play_Seal_Roar", "en_seal_roar.wav", 1.5, "N", "OnAlert", "물개 울음", 3, -6, "wav", "stereo", 44100, "SFX", "", "draft"),
    (23, "Enemy", "Orca_Splash", "EN_M22_ORCA", "Play_Orca_Splash", "en_orca_splash.wav", 1.2, "N", "OnAttack", "범고래 물보라", 3, -6, "wav", "stereo", 44100, "SFX", "", "draft"),
    (24, "Object", "Ice_Crack", "OBJ_M22_ICE_CRACK", "Play_Ice_Crack", "obj_ice_crack.wav", 0.9, "N", "OnCrack", "얼음 갈라짐 시작", 3, -6, "wav", "stereo", 44100, "SFX", "", "final"),
    (25, "Object", "Ice_Break", "OBJ_M22_ICE_BREAK", "Play_Ice_Break", "obj_ice_break.wav", 1.1, "N", "OnBreak", "얼음 붕괴", 2, -6, "wav", "stereo", 44100, "SFX", "", "final"),
    (26, "Object", "Switch_On", "OBJ_M22_SWITCH_ON", "Play_Switch_On", "obj_switch_on.wav", 0.4, "N", "OnSwitchOn", "스위치 온", 5, -6, "wav", "stereo", 44100, "SFX", "", "final"),
    (27, "Object", "Gate_Open", "OBJ_M22_GATE_OPEN", "Play_Gate_Open", "obj_gate_open.wav", 1.5, "N", "OnOpen", "게이트 열림", 4, -6, "wav", "stereo", 44100, "SFX", "", "draft"),
]

VOICE_ROWS = [
    (1, "Narrator", "VO_M22_INTRO_01", "펭귄의 모험이 지금 시작돼요!", "Bright", "타이틀 인트로", 2.4, "vo_intro_01.wav", ""),
    (2, "Narrator", "VO_M22_TUTOR_01", "화살표를 따라 이동해 보세요.", "Neutral", "튜토리얼 이동", 2.2, "vo_tutor_move.wav", ""),
    (3, "Narrator", "VO_M22_TUTOR_02", "점프 버튼을 눌러 장애물을 넘어요!", "Encourage", "튜토리얼 점프", 2.6, "vo_tutor_jump.wav", ""),
    (4, "Narrator", "VO_M22_CLEAR", "성공! 다음 지역으로 출발해요.", "Bright", "스테이지 클리어", 2.0, "vo_clear.wav", ""),
    (5, "Narrator", "VO_M22_FAIL", "괜찮아요, 다시 도전해요!", "Warm", "실패 리트라이", 2.0, "vo_retry.wav", ""),
]

REF_ROWS = [
    ("BGM", "Lo-fi arctic ambience, playful orchestra (YouTube/Ref)", "톤/무드 참고"),
    ("SFX", "Snow footsteps, ice slide, wind blizzard collections", "소스 레퍼런스"),
    ("VO", "Friendly kids narration style", "나레이션 톤 가이드"),
]


def style_header(ws):
    for col_idx, title in enumerate(SHEET_HEADERS[ws.title], start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    # 열 너비
    widths = COL_WIDTHS.get(ws.title)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    # 첫 행 고정 및 필터
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_rows(ws, rows):
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, v in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.border = BORDER
            cell.alignment = LEFT if c_idx in (5, 10, 16, 17) else CENTER


def build_workbook():
    wb = Workbook()

    # 사운드리스트
    ws1 = wb.active
    ws1.title = "사운드리스트"
    style_header(ws1)
    write_rows(ws1, SFX_ROWS)

    # 음성 스크립트
    ws2 = wb.create_sheet("음성 스크립트")
    style_header(ws2)
    write_rows(ws2, VOICE_ROWS)

    # ref
    ws3 = wb.create_sheet("ref")
    style_header(ws3)
    write_rows(ws3, REF_ROWS)

    return wb


def main():
    wb = build_workbook()
    wb.save(FILENAME)
    print(f"✅ 생성 완료: {FILENAME}")


if __name__ == "__main__":
    main()
